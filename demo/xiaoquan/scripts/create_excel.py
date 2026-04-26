import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

os.makedirs("output", exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "人员信息"

headers = ["姓名", "年龄"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="4472C4")
    cell.alignment = Alignment(horizontal="center")

data = [
    ("张三", 28),
    ("李四", 35),
    ("王五", 22),
]
for row_data in data:
    ws.append(row_data)

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 10

wb.save("output/人员信息.xlsx")
print("文件已生成：output/人员信息.xlsx")
