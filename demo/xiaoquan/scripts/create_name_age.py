import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

os.makedirs("output", exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "人员信息"

# 表头样式
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="4472C4")
header_align = Alignment(horizontal="center", vertical="center")

for col, title in enumerate(["姓名", "年龄"], start=1):
    cell = ws.cell(row=1, column=col, value=title)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# 示例数据
rows = [("张三", 28), ("李四", 35), ("王五", 22)]
for row in rows:
    ws.append(row)

# 数据行居中
data_align = Alignment(horizontal="center")
for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2):
    for cell in r:
        cell.font = Font(name="Arial")
        cell.alignment = data_align

ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 10
ws.row_dimensions[1].height = 22

wb.save("output/姓名年龄.xlsx")
print("OK")
