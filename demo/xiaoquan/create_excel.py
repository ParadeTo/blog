from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

wb = Workbook()
sheet = wb.active
sheet.title = "人员信息"

headers = ["姓名", "年龄"]
for col, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col, value=header)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="4472C4")
    cell.alignment = Alignment(horizontal="center")

data = [
    ("张三", 28),
    ("李四", 35),
    ("王五", 22),
]
for row_data in data:
    sheet.append(row_data)

sheet.column_dimensions["A"].width = 20
sheet.column_dimensions["B"].width = 10

output_path = "outputs/人员信息.xlsx"
os.makedirs("outputs", exist_ok=True)
wb.save(output_path)
print(f"saved:{output_path}")
